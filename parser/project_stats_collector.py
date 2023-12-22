import os
import re
import openpyxl
import xlrd
import Levenshtein
import pandas as pd

from datetime import datetime
from fastq_parser import parse_fastq_stats_folder, parse_multiqc_files
from sciebo_parser import parse_sciebo_report

# Predefined dictionary for mapping sequencing kits to expected clusters
sequencing_kit_to_clusters = {
    'nextseq 500/550 high output kit v2.5 (75 cycles)': '400 mio.',
    'nextseq 500/550 high output kit v2.5 (150 cycles)': '400 mio.',
    'nextseq 500/550 mid output kit v2.5 (300 cycles)': '130 mio.',
    'nextseq 500/550 mid output kit v2.5 (150 cycles)': '130 mio.',
    'miseq reagent kit v3 (150-cycle)': '22–25 million',
    'miseq reagent kit v3 (600-cycles)': '22–25 million',
    'miseq reagent kit v2 (50-cycles)': '12-15 million',
    'miseq reagent kit v2 (300-cycles)': '12-15 million',
    'miseq reagent kit v2 (500-cycles)': '12-15 million',
    'miseq reagent micro kit v2 (300-cycles)': '4 million',
    'miseq reagent nano kit v2 (300-cycles)': '1 million',
    'miseq reagent nano kit v2 (500-cycles)': '1 million',
    'novaseq 6000 s4 reagent kit v1.5 (300 cycles)': '8–10 billion',
    'novaseq 6000 s4 reagent kit v1.5 (200 cycles)': '8–10 billion',
    'novaseq 6000 s4 reagent kit v1.5 (35 cycles)': '8–10 billion',
    'novaseq 6000 s2 reagent kit v1.5 (300 cycles)': '3.3–4.1 billion',
    'novaseq 6000 s2 reagent kit v1.5 (200 cycles)': '3.3–4.1 billion',
    'novaseq 6000 s2 reagent kit v1.5 (100 cycles)': '3.3–4.1 billion',
    'novaseq 6000 s1 reagent kit v1.5 (300 cycles)': '1.3–1.6 billion',
    'novaseq 6000 s1 reagent kit v1.5 (200 cycles)': '1.3–1.6 billion',
    'novaseq 6000 s1 reagent kit v1.5 (100 cycles)': '1.3–1.6 billion',
    'novaseq 6000 sp 500 cycles': '650–800 million',
    'novaseq 6000 sp 300 cycles': '650–800 million',
    'novaseq 6000 sp 200 cycles)': '650–800 million',
    'novaseq 6000 sp 100 cycles': '650–800 million',
    }

fastq_folder_path = os.path.join("..", "fastq")
bcl_folder_path = os.path.join("..", "fastq")
sciebo_folder_path = os.path.join("..", "statistics","sciebo","2023")


def main():
    fastq_folders = os.listdir(fastq_folder_path)
    fastq_dates = get_fastq_date(fastq_folders)
    fastq_sequencers = get_sequencer(fastq_folders)
    data = {
        "Project Name": fastq_folders,
        "Date": fastq_dates,
        "Sequencer": fastq_sequencers,
        "std": [None] * len(fastq_folders),
        "cv": [None] * len(fastq_folders),
        "undetermined reads percentage (%)": [None] * len(fastq_folders),
        "most common undetermined barcode": [None] * len(fastq_folders),
        "most common undetermined barcode percentage (%)": [None] * len(fastq_folders),
        "undetermined distribution string": [None] * len(fastq_folders),
        "read distribution": [None] * len(fastq_folders),
        "sequencing_kit": [None] * len(fastq_folders),
        "cycles_read_1": [None] * len(fastq_folders),
        "cycles_index_1": [None] * len(fastq_folders),
        "cycles_read_2": [None] * len(fastq_folders),
        "cycles_index_2": [None] * len(fastq_folders),
        "density": [None] * len(fastq_folders),
        "clusters_pf": [None] * len(fastq_folders),
        "yields (Gb)": [None] * len(fastq_folders),
        "q_30": [None] * len(fastq_folders),
        "project_name": [None] * len(fastq_folders),
        "total read count": [None] * len(fastq_folders),
    }
    
    df = pd.DataFrame(data)
    df['Date'] = pd.to_datetime(df['Date'], format='%d.%m.%Y')
    df = df.sort_values(by="Date")
    df = df.set_index(['Project Name'])

    for folder in fastq_folders:
        # First check that it is a sequencing folder
        if re.match(r'^\d{6}', folder) is None:
            print("The folder is not in the expected fastq project format")
            continue
        df.loc[folder,'std'] , df.loc[folder,'cv'], df.loc[folder,'undetermined reads percentage (%)'], df.loc[folder,'read distribution'], df.loc[folder,'total read count']  = parse_multiqc_files(folder)
        df.loc[folder,'most common undetermined barcode'], df.loc[folder,'undetermined distribution string'], df.loc[folder,'most common undetermined barcode percentage (%)'] =  parse_fastq_stats_folder(folder)
        sciebo_report_path = find_corresponding_sciebo(folder)
        if sciebo_report_path is not None:
            sciebo_report_path = os.path.join(sciebo_folder_path, sciebo_report_path)
            df.loc[folder,'sequencing_kit'], df.loc[folder,'cycles_read_1'], df.loc[folder,'cycles_index_1'], df.loc[folder,'cycles_read_2'], df.loc[folder,'cycles_index_2'], df.loc[folder,'density'], df.loc[folder,'clusters_pf'], df.loc[folder,'yields (Gb)'], df.loc[folder,'q_30'], df.loc[folder,'project_name'] = parse_sciebo_report(sciebo_report_path) 
        # TODO: Add BCL data
            
    df['sequencing_kit'] = df['sequencing_kit'].str.lower()
    df['Expected Clusters'] = df['sequencing_kit'].map(sequencing_kit_to_clusters)
    df['total read count'] = df['total read count'].map(format_with_commas)
 
    return df


def find_corresponding_sciebo(fastq_folder):
    sequence_date_prefix = fastq_folder.split('_')[0]
    flowcell_id = fastq_folder.split('_')[3][1:]
    sciebo_files = os.listdir(sciebo_folder_path)
    sciebo_candidates = [file for file in sciebo_files if file.startswith(sequence_date_prefix)]
    if len(sciebo_candidates) == 0:
        print(f"zero sciebo candidates")
        return None
    elif len(sciebo_candidates) == 1:
        print(f"single sciebo match found! {sciebo_candidates[0]}")
        return sciebo_candidates[0]
    else:
        for sciebo_file in sciebo_candidates: 
            if sciebo_fastq_match(flowcell_id, sciebo_file):
                print(f"sciebo match found from multiple candidates! {sciebo_file}")
                return sciebo_file
    print(f"No sciebo match was found for {fastq_folder}")
    return None


def extract_date_from_folder(folder_name):
    try:
        # Extracting the YYMMDD part from the folder name
        date_str = folder_name[:6]
        # Converting the date string to a datetime object
        date_obj = datetime.strptime(date_str, '%y%m%d')
        # Formatting the datetime object as 'DD.MM.YYYY'
        formatted_date = date_obj.strftime('%d.%m.%Y')
        return formatted_date

    except ValueError:
        # Handle the case where the folder name doesn't match the expected format
        return None


def get_fastq_date(fastq_folders):
    dates = [extract_date_from_folder(folder) for folder in fastq_folders]
    return dates

def get_sequencer(fastq_folders):
    sequencers = [extract_sequencer_from_folder(folder) for folder in fastq_folders]
    return sequencers

def extract_sequencer_from_folder(folder_name):
    name_array = folder_name.split('_')
    if len(name_array) <= 1 :
        return ""
    instrument_id = name_array[1]
    if instrument_id.startswith("NB501289"):
        return "nextseq500"
    elif instrument_id.startswith("M00818"):
        return "miseq1"
    elif instrument_id.startswith("M04404"):
        return "miseq2"
    elif instrument_id.startswith("A01742"):
        return "novaseq"
    else: 
        return ""

def sciebo_fastq_match(flowcell_id, sciebo_file):
    report_path = os.path.join(sciebo_folder_path, sciebo_file)

    if sciebo_file.lower().endswith(".xls"):
        # Open the Excel file
        workbook = xlrd.open_workbook(report_path)
        # Iterate through sheets
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            # Iterate through rows and columns
            for i in range(sheet.nrows):
                for j in range(sheet.ncols):
                    cell_value = sheet.cell_value(i, j)
                    if type(cell_value) is str:
                        threshold = 1
                        distance = Levenshtein.distance(cell_value, cell_value)
                        if distance <= threshold:
                            return True
                        # if flowcell_id in cell_value:
                        #     return True

    elif sciebo_file.lower().endswith(".xlsx"):
        # Open the Excel file
        workbook = openpyxl.load_workbook(filename=report_path, data_only=True)
        # Iterate through sheets
        for excel_sheet_name in workbook.sheetnames:
            sheet = workbook[excel_sheet_name]
            for i in range(1,sheet.max_row):
                for j in range(1, sheet.max_column):
                    cell_value = sheet.cell(row=i, column=j).value
                    if type(cell_value) is str:
                        threshold = 1
                        distance = Levenshtein.distance(cell_value, cell_value)
                        if distance <= threshold:
                            return True
                        # if flowcell_id in cell_value:
                        #     return True
                    
    return False
    
def format_with_commas(number):
    if pd.notna(number):  # Check if the value is not None
        return '{:,}'.format(number)
    else:
        return None

if __name__ == '__main__':
    """
    Create a data frame containing different statistics from all the sequencing projects
    """
    df = main()
    df.to_csv('sequencing_statistics.csv', index=True)
